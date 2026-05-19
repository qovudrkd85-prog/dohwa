# -*- coding: utf-8 -*-
import sys
sys.path.insert(0, r'E:\개인 FIle\01_Projects\02_hydro_work_auto\src\python')
import openpyxl

TPL = r'E:\배평강\2025 project\2025_project\2025_지평지구\02_기초자료\07_실강우량 조사_우량주상도\2026.05.18_우량주상도_re\지평지구_(은척)집중호우시 강우조사(우량주상도) .xlsx'
wb = openpyxl.load_workbook(TPL)
print('SHEETS:', wb.sheetnames)
ws = wb[wb.sheetnames[0]]
print('sheet1 dim:', ws.dimensions, 'charts:', len(ws._charts))
last = None
for r in range(4, 300):
    if ws.cell(row=r, column=1).value is not None:
        last = r
print('sheet1 last data row(A):', last)
for ch in ws._charts:
    try:
        print('CHART anchor:', ch.anchor)
        for s in ch.series:
            print('  series val:', s.val.numRef.f if s.val and s.val.numRef else None,
                  'cat:', s.cat.numRef.f if s.cat and s.cat.numRef else (s.cat.strRef.f if s.cat and s.cat.strRef else None))
    except Exception as e:
        print('  chart err:', e)

# 시강우 파일
from aws_hyetograph.portal_xls_importer import import_portal_xls
SG = r'E:\개인 FIle\00_dohwa\03_토목기술팀\01_data-download-desk\업무\2026-05-18_은척_기왕호우우량주상도\시강우\01_은척_2011.08.08\시강우_20110807-20110809.xls'
ds = import_portal_xls(SG)
recs = ds.records
print('시강우 records:', len(recs))
r0 = recs[0]
print('record attrs:', [a for a in dir(r0) if not a.startswith('_')])
print('first 3:', [(str(getattr(r, "dt", getattr(r, "datetime", "?"))), getattr(r, "rain_mm", "?")) for r in recs[:3]])
nz = [r for r in recs if (getattr(r, 'rain_mm', 0) or 0) > 0]
if nz:
    print('first non-zero:', str(getattr(nz[0], 'dt', getattr(nz[0], 'datetime', '?'))), getattr(nz[0], 'rain_mm', '?'))
