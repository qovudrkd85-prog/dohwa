# -*- coding: utf-8 -*-
"""사상01 검증: 다운로드 시강우 데이터 vs 템플릿 sheet1 기존값."""
import sys
sys.path.insert(0, r'E:\개인 FIle\01_Projects\02_hydro_work_auto\src\python')
import openpyxl
from aws_hyetograph.portal_xls_importer import import_portal_xls

TPL = r'E:\배평강\2025 project\2025_project\2025_지평지구\02_기초자료\07_실강우량 조사_우량주상도\2026.05.18_우량주상도_re\지평지구_(은척)집중호우시 강우조사(우량주상도) .xlsx'
SG = r'E:\개인 FIle\00_dohwa\03_토목기술팀\01_data-download-desk\업무\2026-05-18_은척_기왕호우우량주상도\시강우\01_은척_2011.08.08\시강우_20110807-20110809.xls'

# 다운로드 시강우: 첫 비0 강우부터 43행
recs = import_portal_xls(SG).records
start = next(i for i, r in enumerate(recs) if (r.rain_mm or 0) > 0)
series = recs[start:start + 43]
print(f'다운로드: 시작 {series[0].dt}  ({len(series)}행)')

# 템플릿 sheet1 기존 B4:B46
wb = openpyxl.load_workbook(TPL)
ws = wb[wb.sheetnames[0]]
tpl = [(ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value) for r in range(4, 47)]
print(f'템플릿  : 시작 {tpl[0][0]}  ({len(tpl)}행)')

# 대조
mismatch = []
for i, (rec, (t_dt, t_v)) in enumerate(zip(series, tpl)):
    dv = rec.rain_mm or 0
    tv = t_v or 0
    if abs(dv - tv) > 0.05:
        mismatch.append((i + 4, str(rec.dt), dv, tv))

print(f'\n=== 검증 결과 ===')
print(f'행 수: 다운로드 {len(series)} / 템플릿 {len(tpl)}')
print(f'불일치: {len(mismatch)}건')
for m in mismatch[:10]:
    print(f'  행{m[0]} {m[1]}: 다운={m[2]} 템플릿={m[3]}')
print('PASS — 템플릿 sheet1 = 다운로드 데이터 일치' if not mismatch
      else 'DIFF — 값이 다름, 템플릿 갱신 필요')
print(f'다운로드 43행 합계: {sum((r.rain_mm or 0) for r in series):.1f} mm')
